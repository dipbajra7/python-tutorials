import openpyxl as xl



file_path = 'C:/Users/dipba/Desktop/Book1.xlsx'
workbook = xl.load_workbook(file_path)
worksheet = workbook['Sheet1']
for row in range(2, worksheet.max_row+1):
    cell = worksheet.cell(row,3)
    print(cell.value)
    new_val = 10 + cell.value
    print('new value = ' + str(new_val))
    worksheet.cell(row,4).value = new_val


workbook.save('C:/Users/dipba/Desktop/Book1_copy.xlsx')
 

